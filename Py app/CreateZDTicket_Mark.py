from openpyxl import load_workbook
from zdesk import Zendesk
from pprint import pprint
from getmail import get_mark_mail
import glob
import os


get_mark_mail()

mark_dir = "C:\\Users\\DmitryDmytrenko\\Documents\\Technology\\DD\\Zendesk\\Py app\\Files From Brad Mark\\*.xlsx"
wb = load_workbook(filename=max(glob.glob(mark_dir), key=os.path.getctime))
sheet = wb.worksheets[0]
print(max(glob.glob(mark_dir), key=os.path.getctime))
print(sheet)

reply = int(input("Please enter 1 to continue with the file and tab above "))
if reply == 1:
    print("Will process the file above to create ZD tickets!")
else:
    print("Tickets won't be created")
    exit()

zendesk = Zendesk(zdesk_url='https://envusa.zendesk.com',
                  zdesk_email='dmitry.dmytrenko@envusa.design',
                  zdesk_password='zgXYhMmdxGTT0gMPVOaFfRf5xUNue4UPNsE4MaGs',
                  zdesk_token=True)

# wb = load_workbook(filename='C:\\Users\\DmitryDmytrenko\\Documents\\Technology\\DD\\Zendesk\\Py app\\Files From Brad Mark\\2018-11-05 08-35-31 AM ZD.xlsx')
# sheet = wb['Mark 2018-11-02 ZD']
needed_rows = list(sheet.rows)[1:]


dictionary = {}

for row in needed_rows:
    fcid = row[5].value
    order_id = row[0].value
    order_date = row[1].value
    name = row[2].value
    mail_code = row[4].value
    address = row[6].value.replace("\n", " ")
    requester_email = row[7].value
    requester_phone = row[8].value
    mrchs = row[9].value
    qty = row[10].value
    program = row[11].value
    description = row[12].value
    reason = row[13].value
    comments = row[14].value
    zd_description = str('\n')+str(fcid)+str('\n')+str(name)+str('\n') +\
                  str(mail_code)+str('\n')+str(address)+str('\n')+str(requester_phone)+str('\n')+str(requester_email)+str('\n\n')

    if not dictionary.get(fcid):
        dictionary[fcid] = {}

    dictionary[fcid] = {
            'order_date': set([order_date] if not dictionary[fcid].get('order_date') else list(dictionary[fcid]['order_date']) + [order_date]),
            'order_id': set([order_id] if not dictionary[fcid].get('order_id') else list(dictionary[fcid]['order_id']) + [order_id]),
            'mail_code': set([mail_code] if not dictionary[fcid].get('mail_code') else list(dictionary[fcid]['mail_code']) + [mail_code]),
            'address': set([address] if not dictionary[fcid].get('address') else list(dictionary[fcid]['address']) + [address]),
            'name': set([name] if not dictionary[fcid].get('name') else list(dictionary[fcid]['name']) + [name]),
            'email': set([requester_email] if not dictionary[fcid].get('requester_email') else list(dictionary[fcid]['requester_email']) + [requester_email]),
            'mrchs': set([mrchs] if not dictionary[fcid].get('mrchs') else list(dictionary[fcid]['mrchs']) + [mrchs]),
            'qty': [qty] if not dictionary[fcid].get('qty') else list(dictionary[fcid]['qty']) + [qty],
            'description': set([description] if not dictionary[fcid].get('description') else list(dictionary[fcid]['description']) + [description]),
            'reason': set([reason] if not dictionary[fcid].get('reason') else list(dictionary[fcid]['reason']) + [reason]),
            'comments': set([comments] if not dictionary[fcid].get('comments') else list(dictionary[fcid]['comments']) + [comments]),
            'zd_description': set([zd_description] if not dictionary[fcid].get('zd_description') else list(dictionary[fcid]['zd_description']) + [zd_description])
        }


for fcid, d in dictionary.items():

    new_ticket = {
        'ticket': {
            'requester': {
                'name': ', '.join([str(i) for i in d['name']]) + str('\n\n'),
                'email': ', '.join([str(i) for i in d['email']]) + str('\n\n'),
            },
            'subject': ', '.join([str(i) for i in d['order_id']]),
            'description': ', '.join([str(i) for i in d['order_id']]) + str('\n')
                           + ', '.join([str(i) for i in d['zd_description']]) + str('\n\n')
                           + 'MRCH:'+str('\n')
                           + ', '.join([str(i) for i in d['mrchs']]) + str('\n\n')
                           + 'Qty:'+str('\n')
                           + ', '.join([str(i) for i in d['qty']]) + str('\n\n')
                           + 'Reason:'+str('\n')
                           + ', '.join([str(i) for i in d['reason']]) + str('\n\n')
                           + 'Comments:'+str('\n')
                           + ', \n'.join([str(i) for i in d['comments']]) + str('\n'),
            'tags': ['campaign_report', 'cat_marketing', 'merch_root_cause_na'],
            # 'assignee_id': 116899611531,
            'group_id': 114096002972,
            'type': 'incident',
            'custom_fields': [
                {
                    "id": 360008222992,
                    "name": "FCID",
                    "raw_name": "FCID",
                    "value": fcid,
                    "default": False
                },
                {
                    "id": 360010948532,
                    "name": "Mail Code",
                    "raw_name": "Mail Code",
                    "value": str(dictionary[fcid].get('mail_code')).strip("{''}"),
                    "default": False
                },
                {
                    "id": 360008222972,
                    "name": "Address",
                    "raw_name": "Address",
                    "value": str(dictionary[fcid].get('address')).strip("{''}"),
                    "default": False
                },
                {
                    "id": 360001808731,
                    "name": "Category",
                    "raw_name": "Category",
                    "value": "Marketing",
                    "default": False
                },
                {
                    "id": 114103807612,
                    "name": "Request Origin",
                    "raw_name": "Request Origin",
                    "value": "Campaign Report",
                    "default": False
                },
                {
                    "id": 360002901632,
                    "name": "Root Cause (Merchandising)",
                    "raw_name": "Root Cause (Merchandising)",
                    "value": "N/A",
                    "default": False
                }
            ]
        }
    }

    result = zendesk.ticket_create(data=new_ticket)

    print(result)

    pprint(new_ticket)


